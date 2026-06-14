"""
Endpoint de carga de archivos Excel.
Soporta tres tipos:
  - Mapa de planta (hoja CAL_LOC)
  - Stock de subinventario
  - Producción / reubicación (hoja con nombre PRODUCCION)
"""
import math
import unicodedata
import re
from datetime import datetime, timezone
from io import BytesIO
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from supabase import Client
from core.supabase import get_supabase_admin
from core.deps import get_current_user

router = APIRouter(prefix="/upload", tags=["upload"])

# Columnas fijas del mapa CAL_LOC (0-indexed)
CAL_LOC_HEADER_ROW = 7
CAL_LOC_COL = {"ZONA": 1, "LOC": 2, "FORMATO": 3, "CAPACIDAD": 10, "OCUPADO": 14, "DISPONIBLE": 17, "PCT": 18}


def _norm(v) -> str:
    """Normaliza texto: mayúsculas, sin tildes, sin caracteres especiales."""
    s = unicodedata.normalize("NFD", str(v).strip().upper())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _float(v, default=0.0) -> float:
    try:
        return float(str(v).strip()) if str(v).strip() else default
    except (ValueError, TypeError):
        return default


def _int(v, default=0) -> int:
    return int(_float(v, default))


def _open_wb(data: bytes):
    return load_workbook(BytesIO(data), read_only=True, data_only=True)


# ── Mapa CAL_LOC ─────────────────────────────────────────────────────────────

def _process_mapa(wb) -> list[dict]:
    ws = wb["CAL_LOC"]
    rows = list(ws.iter_rows(values_only=True))
    records = []
    for row in rows[CAL_LOC_HEADER_ROW + 1 :]:
        zona = str(row[CAL_LOC_COL["ZONA"]] or "").strip()
        loc = str(row[CAL_LOC_COL["LOC"]] or "").strip()
        if not zona.startswith("ZONA") or not loc:
            continue
        cap = _int(row[CAL_LOC_COL["CAPACIDAD"]])
        ocup = _int(row[CAL_LOC_COL["OCUPADO"]])
        disp = _int(row[CAL_LOC_COL["DISPONIBLE"]]) or (cap - ocup)
        pct = _float(row[CAL_LOC_COL["PCT"]])
        if pct > 5:
            pct /= 100
        records.append({
            "zona": zona,
            "localizador": loc,
            "formato": str(row[CAL_LOC_COL["FORMATO"]] or "Mezcla").strip(),
            "capacidad": cap,
            "ocupado": ocup,
            "disponible": disp,
            "pct_ocupacion": round(pct * 10000) / 10000,
            "activo": True,
        })
    if not records:
        raise HTTPException(status_code=422, detail="CAL_LOC: no se encontraron filas ZONA válidas")
    return records


async def _upsert_localizadores(db: Client, records: list[dict]) -> int:
    BATCH = 200
    for i in range(0, len(records), BATCH):
        batch = records[i : i + BATCH]
        db.from_("localizadores").upsert(batch, on_conflict="zona,localizador").execute()
    return len(records)


# ── Stock ─────────────────────────────────────────────────────────────────────

def _detect_header(rows: list[tuple]) -> dict:
    """Detecta la fila de encabezados y las columnas relevantes."""
    for i, row in enumerate(rows[:30]):
        cells = [_norm(c) for c in row]
        loc_idx = next(
            (j for j, c in enumerate(cells) if c in ("LOCALIZADOR", "LOC", "LOCALIZACION") or c.startswith("LOCALIZ")),
            -1,
        )
        if loc_idx < 0:
            continue

        def find(keywords, exclude=()) -> int:
            for kw in keywords:
                j = next(
                    (k for k, c in enumerate(cells) if k not in exclude and (c == kw or c.startswith(kw))),
                    -1,
                )
                if j >= 0:
                    return j
            return -1

        subinv_idx = find(("SUBINVENTARIO", "SUBINV", "SUB INVENTARIO", "SUBALMACEN", "SUBIN"))
        formato_idx = find(("FORMATO", "FORMAT", "TIPO"), exclude={loc_idx, subinv_idx} if subinv_idx >= 0 else {loc_idx})
        lote_idx = find(("LOTE", "LOT", "NUM LOTE", "NRO LOTE"), exclude={loc_idx, subinv_idx, formato_idx})
        tarima_idx = find(("TARIMA", "TARIMAS", "PALLET", "PALLETS", "ESTIBA", "PALETA"),
                         exclude={loc_idx, subinv_idx})
        cajas_idx = find(("CAJAS", "CAJA", "BOXES", "FRACCION"),
                        exclude={loc_idx, subinv_idx, tarima_idx})
        cant_idx = find(("CANTIDAD", "CANT", "QTY", "STOCK", "SALDO", "TOTAL", "UNIDADES"),
                       exclude={loc_idx, subinv_idx, tarima_idx, cajas_idx})

        return {
            "header_row": i,
            "loc": loc_idx,
            "subinv": subinv_idx,
            "formato": formato_idx,
            "lote": lote_idx,
            "tarima": tarima_idx,
            "cajas": cajas_idx,
            "cant": cant_idx,
        }
    raise HTTPException(status_code=422, detail="No se encontró columna LOCALIZADOR en las primeras 30 filas")


def _process_stock(wb) -> tuple[dict[str, int], dict[str, dict[str, int]], list[dict]]:
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    col = _detect_header(rows)

    total_by_loc: dict[str, int] = {}
    by_subinv: dict[str, dict[str, int]] = {}
    subinv_items: list[dict] = []

    for row in rows[col["header_row"] + 1 :]:
        raw = str(row[col["loc"]] or "").strip()
        if not raw:
            continue
        loc = raw.upper()

        tarima_val = _float(row[col["tarima"]]) if col["tarima"] >= 0 else 0
        cajas_val = _float(row[col["cajas"]]) if col["cajas"] >= 0 else 0
        if col["tarima"] >= 0:
            pallets = int(tarima_val.__ceil__()) + (1 if cajas_val > 0 else 0)
        else:
            pallets = 1

        total_by_loc[loc] = total_by_loc.get(loc, 0) + pallets

        if col["subinv"] >= 0:
            si = str(row[col["subinv"]] or "").strip()
            fmt = str(row[col["formato"]] or "").strip() if col["formato"] >= 0 else ""
            lot = str(row[col["lote"]] or "").strip() if col["lote"] >= 0 else ""
            if si:
                if si not in by_subinv:
                    by_subinv[si] = {}
                by_subinv[si][loc] = by_subinv[si].get(loc, 0) + pallets
                subinv_items.append({"loc": loc, "subinv": si, "formato": fmt, "lote": lot, "pallets": pallets})

    return total_by_loc, by_subinv, subinv_items


async def _apply_stock(db: Client, total_by_loc: dict[str, int]) -> int:
    resp = db.from_("localizadores").select("zona,localizador,capacidad").execute()
    db_map = {r["localizador"].strip().upper(): r for r in (resp.data or [])}

    updates = []
    for loc, info in db_map.items():
        qty = total_by_loc.get(loc, 0)
        ocupado = min(2_147_483_647, max(0, round(qty)))
        disponible = max(0, info["capacidad"] - ocupado)
        raw_pct = ocupado / info["capacidad"] if info["capacidad"] > 0 else 0
        pct = min(9.9999, round(raw_pct * 10000) / 10000)
        updates.append({"zona": info["zona"], "localizador": loc, "ocupado": ocupado,
                        "disponible": disponible, "pct_ocupacion": pct})

    BATCH = 200
    for i in range(0, len(updates), BATCH):
        db.from_("localizadores").upsert(updates[i : i + BATCH], on_conflict="zona,localizador").execute()

    return sum(1 for loc in db_map if total_by_loc.get(loc, 0) > 0)


# ── Producción / Reubicación ──────────────────────────────────────────────────

def _process_produccion(wb) -> list[dict]:
    sheet_name = next((n for n in wb.sheetnames if "PRODUCCION" in _norm(n)), None)
    if sheet_name is None:
        raise HTTPException(status_code=422, detail="No se encontró hoja PRODUCCION")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Detectar fila de encabezados buscando DESCRIP + COD
    header_row = -1
    col_map: dict[str, int] = {}
    for i, row in enumerate(rows[:15]):
        cells = [_norm(c) for c in row]
        has_desc = any("DESCRIP" in c for c in cells)
        has_cod = any("COD" in c or "CODIGO" in c for c in cells)
        if not has_desc or not has_cod:
            continue
        header_row = i
        for idx, c in enumerate(cells):
            if "COD" in c and ("ORG" in c or ("INV" in c and len(c) > 8)):
                col_map.setdefault("cod_org_inv", idx)
            elif c == "CODIGO" or (c.startswith("COD") and len(c) <= 6):
                col_map.setdefault("codigo", idx)
            elif "DESCRIP" in c:
                col_map.setdefault("descripcion", idx)
            elif "SUBIN" in c and "ORIG" in c:
                col_map.setdefault("subinventario_origen", idx)
            elif "LOCAL" in c and "ORIG" in c:
                col_map.setdefault("localizador_origen", idx)
            elif c == "LOTE":
                col_map.setdefault("lote", idx)
            elif ("CAN" in c or "CANT" in c) and "FIS" in c:
                col_map.setdefault("cantidad_fisica", idx)
            elif c in ("PALLETS", "TARIMAS", "PLT") or c.startswith("PALLET"):
                col_map.setdefault("pallets", idx)
            elif c == "CAJAS":
                col_map.setdefault("cajas", idx)
            elif "SUBIN" in c and "DEST" in c:
                col_map.setdefault("subinventario_destino", idx)
            elif "LOCAL" in c and "DEST" in c:
                col_map.setdefault("localizador_destino", idx)
            elif "RESPON" in c:
                col_map.setdefault("responsable", idx)
            elif c == "CONTEO" or c.startswith("CONTEO"):
                col_map.setdefault("conteo", idx)
        break

    if header_row < 0 or "descripcion" not in col_map:
        raise HTTPException(status_code=422, detail="No se encontró fila de encabezados válida en hoja PRODUCCION")

    def get(row, key):
        idx = col_map.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    raw_rows = []
    for row in rows[header_row + 1 :]:
        desc = str(get(row, "descripcion") or "").strip()
        if not desc:
            continue
        raw_rows.append({
            "cod_org_inv": str(get(row, "cod_org_inv") or "").strip() or None,
            "codigo": str(get(row, "codigo") or "").strip() or None,
            "descripcion": desc,
            "subinventario_origen": str(get(row, "subinventario_origen") or "").strip() or None,
            "localizador_origen": str(get(row, "localizador_origen") or "").strip() or None,
            "lote": str(get(row, "lote") or "").strip() or None,
            "cantidad_fisica": _float(get(row, "cantidad_fisica")),
            "pallets": _int(get(row, "pallets")),
            "cajas": _int(get(row, "cajas")),
            "subinventario_destino": str(get(row, "subinventario_destino") or "").strip() or None,
            "localizador_destino": str(get(row, "localizador_destino") or "").strip() or None,
            "responsable": str(get(row, "responsable") or "").strip() or None,
            "conteo": _int(get(row, "conteo")) if "conteo" in col_map else None,
        })

    if not raw_rows:
        raise HTTPException(status_code=422, detail="PRODUCCION: no se encontraron filas con datos")
    return raw_rows


async def _insert_produccion(db: Client, raw_rows: list[dict]) -> int:
    # Calcular inv_pe: ocupado del destino + pallets de la fila
    dest_locs = list({r["localizador_destino"].upper() for r in raw_rows if r["localizador_destino"]})
    loc_resp = db.from_("localizadores").select("localizador,ocupado").in_("localizador", dest_locs).execute()
    ocupado_map = {r["localizador"].strip().upper(): r["ocupado"] or 0 for r in (loc_resp.data or [])}

    # Resolver cajas_por_pallet del catálogo para filas con pallets=0 pero cajas>0
    codigos = list({r["codigo"] for r in raw_rows if r.get("codigo")})
    cat_map: dict[str, float] = {}
    if codigos:
        cat_resp = db.from_("catalogo_productos").select("codigo,cajas_por_pallet").in_("codigo", codigos).execute()
        for c in (cat_resp.data or []):
            if c.get("cajas_por_pallet"):
                cat_map[c["codigo"]] = float(c["cajas_por_pallet"])

    now = datetime.now(timezone.utc).isoformat()
    inserts = []
    for r in raw_rows:
        loc_key = (r["localizador_destino"] or "").upper()
        pallets = r.get("pallets") or 0
        cajas = r.get("cajas") or 0
        if pallets == 0 and cajas > 0:
            cpp = cat_map.get(r.get("codigo") or "", 0)
            pallets = math.ceil(cajas / cpp) if cpp > 0 else 1
        inv_pe = ocupado_map.get(loc_key, 0) + pallets
        inserts.append({**r, "inv_pe": inv_pe, "estado": "pendiente", "created_at": now, "updated_at": now})

    BATCH = 50
    for i in range(0, len(inserts), BATCH):
        db.from_("lineas_reubicacion").insert(inserts[i : i + BATCH]).execute()

    return len(inserts)


# ── Preview / Validación ──────────────────────────────────────────────────────

@router.post("/preview")
async def preview_excel(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
    db: Client = Depends(get_supabase_admin),
):
    """
    Analiza el archivo sin escribir en BD.

    - Mapa CAL_LOC   → devuelve conteo y formatos
    - Producción     → enriquece filas con zona + formato del localizador destino
    - Stock          → valida compatibilidad de formatos archivo vs BD
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx o .xls")

    data = await file.read()
    try:
        wb = _open_wb(data)
    except Exception:
        raise HTTPException(status_code=422, detail="No se pudo leer el archivo Excel")

    sheet_names = wb.sheetnames

    # ── Mapa ─────────────────────────────────────────────────────────────────
    if "CAL_LOC" in sheet_names:
        records = _process_mapa(wb)
        formatos = list({str(r.get("formato", "")) for r in records if r.get("formato")})
        return {"type": "mapa", "count": len(records), "formatos": formatos}

    # ── Producción ────────────────────────────────────────────────────────────
    if any("PRODUCCION" in _norm(n) for n in sheet_names):
        raw_rows = _process_produccion(wb)

        dest_locs = list({r["localizador_destino"].upper() for r in raw_rows if r["localizador_destino"]})
        loc_resp = db.from_("localizadores").select("zona,localizador,formato").in_("localizador", dest_locs).execute()
        loc_map = {r["localizador"].strip().upper(): r for r in (loc_resp.data or [])}

        rows_out = []
        locs_not_found: list[str] = []
        for i, r in enumerate(raw_rows):
            loc_key = (r["localizador_destino"] or "").upper()
            db_loc = loc_map.get(loc_key)
            rows_out.append({
                "_id": i,
                **r,
                "zona_destino": db_loc["zona"] if db_loc else "",
                "formato_destino": (db_loc.get("formato") or "") if db_loc else "",
                "loc_encontrado": db_loc is not None,
            })
            if loc_key and not db_loc and loc_key not in locs_not_found:
                locs_not_found.append(loc_key)

        return {"type": "produccion", "rows": rows_out, "locs_not_found": locs_not_found, "total": len(rows_out)}

    # ── Stock ─────────────────────────────────────────────────────────────────
    ws = wb.active
    raw = list(ws.iter_rows(values_only=True))
    col = _detect_header(raw)   # raises 422 si no hay LOCALIZADOR

    # Localizadores de BD con formato
    loc_resp = db.from_("localizadores").select("zona,localizador,formato").execute()
    db_loc_map = {r["localizador"].strip().upper(): r for r in (loc_resp.data or [])}

    # Leer (localizador → formato) del archivo  (última ocurrencia si hay varias)
    file_locs: dict[str, str] = {}
    for row in raw[col["header_row"] + 1:]:
        raw_loc = str(row[col["loc"]] or "").strip()
        if not raw_loc:
            continue
        loc_key = raw_loc.upper()
        fmt = str(row[col["formato"]] or "").strip() if col["formato"] >= 0 else ""
        file_locs[loc_key] = fmt

    has_format_col = col["formato"] >= 0

    # Construir format_checks
    format_checks = []
    for loc_key, fmt_archivo in file_locs.items():
        db_loc = db_loc_map.get(loc_key)
        if db_loc:
            fmt_db = (db_loc.get("formato") or "").strip()
            comodin = not fmt_db or fmt_db.upper() in ("MEZCLA", "MIX", "VACIO")
            match = not fmt_archivo or comodin or fmt_archivo.upper() == fmt_db.upper()
            format_checks.append({
                "localizador": loc_key,
                "zona": db_loc["zona"],
                "formato_archivo": fmt_archivo,
                "formato_db": fmt_db,
                "match": match,
                "en_db": True,
            })
        else:
            format_checks.append({
                "localizador": loc_key,
                "zona": "",
                "formato_archivo": fmt_archivo,
                "formato_db": "",
                "match": False,
                "en_db": False,
            })

    mismatches = sum(1 for c in format_checks if not c["match"] and c["en_db"])
    not_in_db = sum(1 for c in format_checks if not c["en_db"])

    return {
        "type": "stock",
        "has_format_col": has_format_col,
        "total_locs": len(file_locs),
        "locs_in_db": sum(1 for k in file_locs if k in db_loc_map),
        "format_checks": format_checks,
        "mismatches_count": mismatches,
        "not_in_db_count": not_in_db,
    }


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post("/excel")
async def upload_excel(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
    db: Client = Depends(get_supabase_admin),
):
    """
    Detecta automáticamente el tipo de archivo y lo procesa:
    - Hoja CAL_LOC → actualiza mapa de planta
    - Hoja PRODUCCION → crea líneas de reubicación pendientes
    - Cualquier otro → actualiza stock de localizadores
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx o .xls")

    data = await file.read()
    try:
        wb = _open_wb(data)
    except Exception:
        raise HTTPException(status_code=422, detail="No se pudo leer el archivo Excel")

    sheet_names = wb.sheetnames

    # Mapa CAL_LOC
    if "CAL_LOC" in sheet_names:
        records = _process_mapa(wb)
        count = await _upsert_localizadores(db, records)
        return {"ok": True, "type": "mapa", "count": count}

    # Producción
    if any("PRODUCCION" in _norm(n) for n in sheet_names):
        raw_rows = _process_produccion(wb)
        count = await _insert_produccion(db, raw_rows)
        return {"ok": True, "type": "produccion", "count": count}

    # Stock
    total_by_loc, by_subinv, subinv_items = _process_stock(wb)
    if not total_by_loc:
        raise HTTPException(status_code=422, detail="No se encontraron filas con localizador válido")
    count = await _apply_stock(db, total_by_loc)
    return {
        "ok": True,
        "type": "stock",
        "count": count,
        "subinventarios": list(by_subinv.keys()),
        "subinv_items": subinv_items,
    }
